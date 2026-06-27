from __future__ import annotations

import argparse
import fnmatch
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_PATTERNS = [
    r"\b[0-9]{3}-[0-9]{4}-[0-9]{4}\b",
    r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
]

DEFAULT_EXCLUDE_PATTERNS = [
    ".git/**",
    ".hg/**",
    ".svn/**",
    ".venv/**",
    "venv/**",
    "__pycache__/**",
    "**/__pycache__/**",
    ".pytest_cache/**",
    ".mypy_cache/**",
    ".ruff_cache/**",
    "build/**",
    "dist/**",
    "*.egg-info/**",
    "**/*.egg-info/**",
]


@dataclass
class Finding:
    path: str
    location: str
    rule: str
    value: str


def main() -> int:
    parser = argparse.ArgumentParser(description="Fail when configured sensitive tokens remain.")
    parser.add_argument("paths", nargs="+", type=Path, help="Files or directories to scan")
    parser.add_argument("--denylist", type=Path, help="JSON file with tokens and regex_patterns")
    parser.add_argument(
        "--exclude",
        action="append",
        default=[],
        metavar="GLOB",
        help="Glob to exclude from directory scans. Can be repeated.",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON findings")
    args = parser.parse_args()

    tokens, patterns = load_denylist(args.denylist)
    findings = list(scan_paths(args.paths, tokens, patterns, exclude_patterns=args.exclude))
    if args.json:
        print(json.dumps([finding.__dict__ for finding in findings], ensure_ascii=False, indent=2))
    else:
        for finding in findings:
            print(f"{finding.path}:{finding.location}: {finding.rule}: {finding.value}")
    return 1 if findings else 0


def load_denylist(path: Path | None) -> tuple[list[str], list[re.Pattern[str]]]:
    tokens: list[str] = []
    pattern_texts = list(DEFAULT_PATTERNS)
    if path is not None:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            raise ValueError("Denylist must be a JSON object.")
        tokens = [str(item) for item in payload.get("tokens", [])]
        pattern_texts.extend(str(item) for item in payload.get("regex_patterns", []))
    return tokens, [re.compile(pattern) for pattern in pattern_texts]


def scan_paths(
    paths: Iterable[Path],
    tokens: list[str],
    patterns: list[re.Pattern[str]],
    exclude_patterns: Iterable[str] | None = None,
) -> Iterable[Finding]:
    excludes = [*DEFAULT_EXCLUDE_PATTERNS, *(exclude_patterns or [])]
    for path in paths:
        resolved = path.expanduser().resolve()
        if resolved.is_dir():
            for file_path in resolved.rglob("*"):
                if file_path.is_file() and not matches_exclude(file_path, resolved, excludes):
                    yield from scan_file(file_path, tokens, patterns)
        elif resolved.is_file() and not matches_exclude(resolved, resolved.parent, excludes):
            yield from scan_file(resolved, tokens, patterns)


def matches_exclude(path: Path, root: Path, patterns: Iterable[str]) -> bool:
    try:
        relative_path = path.relative_to(root)
    except ValueError:
        relative_path = path
    relative_text = relative_path.as_posix()
    name = path.name
    for pattern in patterns:
        normalized = pattern.replace("\\", "/")
        if fnmatch.fnmatch(relative_text, normalized) or fnmatch.fnmatch(name, normalized):
            return True
    return False


def scan_file(path: Path, tokens: list[str], patterns: list[re.Pattern[str]]) -> Iterable[Finding]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        yield from scan_workbook(path, tokens, patterns)
        return
    if path.suffix.lower() in {".md", ".txt", ".json", ".toml", ".yml", ".yaml", ".py"}:
        text = path.read_text(encoding="utf-8", errors="ignore")
        yield from scan_text(str(path), "text", text, tokens, patterns)


def scan_workbook(
    path: Path,
    tokens: list[str],
    patterns: list[re.Pattern[str]],
) -> Iterable[Finding]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    for sheet in workbook.worksheets:
        yield from scan_text(str(path), f"sheet:{sheet.title}", sheet.title, tokens, patterns)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    yield from scan_text(
                        str(path),
                        f"{sheet.title}!{cell.coordinate}",
                        cell.value,
                        tokens,
                        patterns,
                    )


def scan_text(
    path: str,
    location: str,
    text: str,
    tokens: list[str],
    patterns: list[re.Pattern[str]],
) -> Iterable[Finding]:
    for token in tokens:
        if token and token in text:
            yield Finding(path=path, location=location, rule="token", value=token)
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            yield Finding(path=path, location=location, rule=pattern.pattern, value=match.group(0))


if __name__ == "__main__":
    raise SystemExit(main())
