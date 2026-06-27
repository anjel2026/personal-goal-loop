from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.workbook import Workbook


@dataclass
class Rules:
    exact_text_replacements: dict[str, str] = field(default_factory=dict)
    regex_replacements: list[tuple[re.Pattern[str], str]] = field(default_factory=list)
    sheet_cell_replacements: dict[str, dict[str, Any]] = field(default_factory=dict)
    preserve_sheet_titles: list[str] = field(default_factory=list)
    auto_rename_sheet_titles: bool = False
    sheet_title_prefix: str = "P"
    sheet_title_width: int = 3
    sheet_title_replacements: dict[str, str] = field(default_factory=dict)
    derive_text_replacements_from_sheet_titles: bool = True
    display_name_cells: list[str] = field(default_factory=lambda: ["D2"])
    clear_comments: bool = True
    clear_hyperlinks: bool = True
    sanitize_properties: bool = True
    metadata_overrides: dict[str, str] = field(default_factory=dict)


@dataclass
class RunSummary:
    source_path: str
    output_path: str
    rules_path: str | None
    dry_run: bool = False
    source_sha256: str | None = None
    output_sha256: str | None = None
    rules_sha256: str | None = None
    policy_version: str = "workbook-anonymizer/v1"
    total_cells_changed: int = 0
    total_comments_removed: int = 0
    total_hyperlinks_removed: int = 0
    derived_text_replacements_count: int = 0
    renamed_sheets: dict[str, str] = field(default_factory=dict)
    sheets: dict[str, dict[str, int]] = field(default_factory=dict)

    def to_dict(self, *, include_sensitive: bool = False) -> dict[str, Any]:
        renamed_sheets: dict[str, str]
        sheets: dict[str, dict[str, int]]
        if include_sensitive:
            source_path = self.source_path
            output_path = self.output_path
            rules_path = self.rules_path
            renamed_sheets = self.renamed_sheets
            sheets = self.sheets
            sensitive_fields_redacted = False
        else:
            source_path = "[redacted]"
            output_path = "[redacted]"
            rules_path = "[redacted]" if self.rules_path is not None else None
            renamed_sheets = _redact_string_map_keys(self.renamed_sheets, "source_sheet")
            sheets = _redact_sheet_summary_keys(self.sheets)
            sensitive_fields_redacted = True

        return {
            "source_path": source_path,
            "output_path": output_path,
            "rules_path": rules_path,
            "dry_run": self.dry_run,
            "source_sha256": self.source_sha256,
            "output_sha256": self.output_sha256,
            "rules_sha256": self.rules_sha256,
            "policy_version": self.policy_version,
            "sensitive_fields_redacted": sensitive_fields_redacted,
            "total_cells_changed": self.total_cells_changed,
            "total_comments_removed": self.total_comments_removed,
            "total_hyperlinks_removed": self.total_hyperlinks_removed,
            "derived_text_replacements_count": self.derived_text_replacements_count,
            "renamed_sheets_count": len(self.renamed_sheets),
            "renamed_sheets": renamed_sheets,
            "sheets": sheets,
        }


def load_rules(path: Path | None) -> Rules:
    if path is None:
        return Rules()

    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Rules file must contain a JSON object.")

    width = int(payload.get("sheet_title_width", 3))
    if width < 1:
        raise ValueError("sheet_title_width must be at least 1.")

    return Rules(
        exact_text_replacements=_to_string_map(payload.get("exact_text_replacements", {})),
        regex_replacements=_compile_regex_rules(payload.get("regex_replacements", [])),
        sheet_cell_replacements=_to_nested_map(payload.get("sheet_cell_replacements", {})),
        preserve_sheet_titles=_to_string_list(payload.get("preserve_sheet_titles", [])),
        auto_rename_sheet_titles=bool(payload.get("auto_rename_sheet_titles", False)),
        sheet_title_prefix=str(payload.get("sheet_title_prefix", "P")),
        sheet_title_width=width,
        sheet_title_replacements=_to_string_map(payload.get("sheet_title_replacements", {})),
        derive_text_replacements_from_sheet_titles=bool(
            payload.get("derive_text_replacements_from_sheet_titles", True)
        ),
        display_name_cells=_to_string_list(payload.get("display_name_cells", ["D2"])),
        clear_comments=bool(payload.get("clear_comments", True)),
        clear_hyperlinks=bool(payload.get("clear_hyperlinks", True)),
        sanitize_properties=bool(payload.get("sanitize_properties", True)),
        metadata_overrides=_to_string_map(payload.get("metadata_overrides", {})),
    )


def anonymize_workbook(
    source_path: Path,
    output_path: Path,
    rules: Rules,
    *,
    dry_run: bool = False,
    rules_path: Path | None = None,
) -> RunSummary:
    keep_vba = source_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_path, keep_vba=keep_vba)
    summary = RunSummary(
        source_path=str(source_path),
        output_path=str(output_path),
        rules_path=str(rules_path) if rules_path else None,
        dry_run=dry_run,
        source_sha256=sha256_file(source_path),
        rules_sha256=sha256_file(rules_path) if rules_path else None,
    )

    sheet_title_map, derived_text_replacements, derived_count = _build_anonymization_maps(
        workbook, rules
    )
    summary.renamed_sheets = dict(sheet_title_map)
    summary.derived_text_replacements_count = derived_count

    exact_replacements = dict(rules.exact_text_replacements)
    for source, replacement in derived_text_replacements.items():
        exact_replacements.setdefault(source, replacement)

    for sheet in workbook.worksheets:
        original_title = sheet.title
        sheet_summary = {
            "cells_changed": 0,
            "comments_removed": 0,
            "hyperlinks_removed": 0,
        }
        cell_replacements = dict(rules.sheet_cell_replacements.get("*", {}))
        cell_replacements.update(rules.sheet_cell_replacements.get(original_title, {}))

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                original_value = cell.value
                changed = False

                if cell.coordinate in cell_replacements:
                    new_value = cell_replacements[cell.coordinate]
                elif _is_formula(original_value):
                    new_value = rewrite_formula_references(original_value, sheet_title_map)
                elif isinstance(original_value, str):
                    new_value = apply_value_rules(
                        original_value,
                        exact_replacements,
                        rules.regex_replacements,
                    )
                else:
                    new_value = original_value

                if new_value != original_value:
                    cell.value = new_value
                    changed = True

                if changed:
                    sheet_summary["cells_changed"] += 1
                    summary.total_cells_changed += 1

                if rules.clear_comments and cell.comment is not None:
                    cell.comment = None
                    sheet_summary["comments_removed"] += 1
                    summary.total_comments_removed += 1

                if rules.clear_hyperlinks and cell.hyperlink is not None:
                    cell.hyperlink = None
                    sheet_summary["hyperlinks_removed"] += 1
                    summary.total_hyperlinks_removed += 1

        summary.sheets[original_title] = sheet_summary

    rename_worksheets(workbook, sheet_title_map)
    if rules.sanitize_properties:
        sanitize_workbook_properties(workbook, rules.metadata_overrides)

    if not dry_run:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        summary.output_sha256 = sha256_file(output_path)
    return summary


def write_report(
    summary: RunSummary,
    report_path: Path,
    *,
    include_sensitive: bool = False,
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(summary.to_dict(include_sensitive=include_sensitive), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def default_output_path(input_path: Path) -> Path:
    suffix = input_path.suffix or ".xlsx"
    return input_path.with_name(f"{input_path.stem}_anonymized{suffix}")


def default_report_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}.anonymize-report.json")


def sha256_file(path: Path | None) -> str | None:
    if path is None:
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _redact_string_map_keys(values: dict[str, str], prefix: str) -> dict[str, str]:
    return {f"{prefix}_{index:03d}": value for index, value in enumerate(values.values(), start=1)}


def _redact_sheet_summary_keys(
    values: dict[str, dict[str, int]],
) -> dict[str, dict[str, int]]:
    return {
        f"sheet_{index:03d}": dict(value)
        for index, value in enumerate(values.values(), start=1)
    }


def apply_value_rules(
    value: str,
    exact_text_replacements: dict[str, str],
    regex_replacements: list[tuple[re.Pattern[str], str]],
) -> str:
    updated = exact_text_replacements.get(value, value)
    for pattern, replacement in regex_replacements:
        updated = pattern.sub(replacement, updated)
    return updated


def rewrite_formula_references(formula: str, sheet_title_map: dict[str, str]) -> str:
    updated = formula
    for original_title, new_title in sorted(
        sheet_title_map.items(), key=lambda item: len(item[0]), reverse=True
    ):
        escaped_original = re.escape(original_title.replace("'", "''"))
        quoted = re.compile(rf"(?<![A-Za-z0-9_])'{escaped_original}'!")
        unquoted = re.compile(rf"(?<![A-Za-z0-9_']){re.escape(original_title)}!")
        updated = quoted.sub(f"'{new_title}'!", updated)
        updated = unquoted.sub(f"{new_title}!", updated)
    return updated


def rename_worksheets(workbook: Workbook, rename_map: dict[str, str]) -> None:
    temp_titles: dict[str, str] = {}
    for index, sheet in enumerate(workbook.worksheets):
        original_title = sheet.title
        if original_title not in rename_map:
            continue
        temp_title = f"__pgl_tmp_{index}__"
        temp_titles[temp_title] = rename_map[original_title]
        sheet.title = temp_title

    for sheet in workbook.worksheets:
        if sheet.title in temp_titles:
            sheet.title = temp_titles[sheet.title]


def sanitize_workbook_properties(workbook: Workbook, overrides: dict[str, str]) -> None:
    props = workbook.properties
    for field_name in [
        "creator",
        "lastModifiedBy",
        "title",
        "subject",
        "description",
        "keywords",
        "category",
        "manager",
        "company",
        "language",
        "version",
        "contentStatus",
    ]:
        if hasattr(props, field_name):
            setattr(props, field_name, overrides.get(field_name, ""))


def _build_anonymization_maps(
    workbook: Workbook, rules: Rules
) -> tuple[dict[str, str], dict[str, str], int]:
    preserve_titles = set(rules.preserve_sheet_titles)
    rename_map: dict[str, str] = {}
    derived_text_replacements: dict[str, str] = dict(rules.exact_text_replacements)
    person_id_by_key: dict[str, str] = {}
    person_occurrence_by_key: dict[str, int] = {}
    derived_added_count = 0
    next_person_index = 1

    def allocate_person_id() -> str:
        nonlocal next_person_index
        if not rules.sheet_title_prefix:
            raise ValueError("sheet_title_prefix must not be empty.")
        person_id = f"{rules.sheet_title_prefix}{next_person_index:0{rules.sheet_title_width}d}"
        next_person_index += 1
        if len(person_id) > 31:
            raise ValueError(f"Generated sheet title exceeds Excel's 31 character limit: {person_id}")
        return person_id

    def add_variants(source_text: str, replacement: str) -> None:
        nonlocal derived_added_count
        for variant in _text_variants(source_text):
            current = derived_text_replacements.get(variant)
            if current is None:
                derived_text_replacements[variant] = replacement
                derived_added_count += 1
            elif current != replacement:
                raise ValueError(
                    f"Conflicting anonymization target for text {source_text!r}: "
                    f"{current!r} vs {replacement!r}"
                )

    for sheet in workbook.worksheets:
        original_title = sheet.title
        if original_title in preserve_titles:
            continue

        explicit_title = rules.sheet_title_replacements.get(original_title)
        if not rules.auto_rename_sheet_titles and explicit_title is None:
            continue

        display_name = _read_display_name(sheet, rules.display_name_cells)
        person_id: str | None = None
        occurrence = 1

        if rules.auto_rename_sheet_titles:
            name_key = _normalize_name_key(display_name or original_title)
            if name_key not in person_id_by_key:
                person_id_by_key[name_key] = allocate_person_id()
            person_id = person_id_by_key[name_key]
            occurrence = person_occurrence_by_key.get(name_key, 0) + 1
            person_occurrence_by_key[name_key] = occurrence

        new_title = explicit_title or (person_id if occurrence == 1 else f"{person_id}_{occurrence}")
        if not new_title:
            continue
        _validate_sheet_title(original_title, new_title, rename_map)
        rename_map[original_title] = new_title

        if person_id and rules.derive_text_replacements_from_sheet_titles:
            add_variants(original_title, person_id)
            if display_name:
                add_variants(display_name, person_id)

    _validate_rename_targets(workbook, rename_map, preserve_titles)
    return rename_map, derived_text_replacements, derived_added_count


def _read_display_name(sheet: Any, coordinates: list[str]) -> str | None:
    for coordinate in coordinates:
        value = sheet[coordinate].value
        if isinstance(value, str) and value and not value.startswith("="):
            return value
    return None


def _validate_sheet_title(
    original_title: str, new_title: str, rename_map: dict[str, str]
) -> None:
    if len(new_title) > 31:
        raise ValueError(
            f"Generated sheet title for {original_title!r} exceeds Excel's 31 character limit: "
            f"{new_title}"
        )
    if new_title in rename_map.values():
        raise ValueError(f"Duplicate anonymized sheet title requested: {new_title}")


def _validate_rename_targets(
    workbook: Workbook, rename_map: dict[str, str], preserve_titles: set[str]
) -> None:
    unrenamed_titles = {sheet.title for sheet in workbook.worksheets if sheet.title not in rename_map}
    for source_title, target_title in rename_map.items():
        if source_title in preserve_titles:
            raise ValueError(f"Sheet {source_title!r} is preserved and cannot be renamed.")
        if target_title in unrenamed_titles:
            raise ValueError(
                f"Anonymized sheet title {target_title!r} collides with an unrenamed sheet title."
            )


def _to_string_map(value: Any) -> dict[str, str]:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("Expected a JSON object.")
    return {str(key): "" if item is None else str(item) for key, item in value.items()}


def _to_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("Expected a JSON array.")
    return [str(item) for item in value]


def _to_nested_map(value: Any) -> dict[str, dict[str, Any]]:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("Expected a JSON object for sheet_cell_replacements.")
    result: dict[str, dict[str, Any]] = {}
    for sheet_name, cells in value.items():
        if not isinstance(cells, dict):
            raise ValueError(
                f"Sheet replacement entry for {sheet_name!r} must be a JSON object."
            )
        result[str(sheet_name)] = {str(coord): replacement for coord, replacement in cells.items()}
    return result


def _compile_regex_rules(value: Any) -> list[tuple[re.Pattern[str], str]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("Expected a JSON array for regex_replacements.")

    compiled: list[tuple[re.Pattern[str], str]] = []
    for index, item in enumerate(value):
        if not isinstance(item, dict):
            raise ValueError(f"Regex replacement #{index + 1} must be a JSON object.")
        pattern = item.get("pattern")
        replacement = item.get("replacement", "")
        if not isinstance(pattern, str):
            raise ValueError(f"Regex replacement #{index + 1} is missing a string pattern.")
        compiled.append((re.compile(pattern), "" if replacement is None else str(replacement)))
    return compiled


def _normalize_name_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    normalized = re.sub(r"\s+", "", normalized)
    return normalized.casefold()


def _text_variants(value: str) -> set[str]:
    normalized = unicodedata.normalize("NFKC", value)
    collapsed = re.sub(r"\s+", "", normalized)
    ascii_spaced = re.sub(r"\s+", " ", normalized).strip()
    fullwidth_spaced = re.sub(r"\s+", "\u3000", normalized).strip()
    return {
        value.strip(),
        normalized.strip(),
        collapsed,
        ascii_spaced,
        fullwidth_spaced,
    }


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")
