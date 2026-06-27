from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment


def make_synthetic_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Name"
    summary["B1"] = "Hours"
    summary["A2"] = "=Alice!D2"
    summary["B2"] = "=Alice!I42"
    summary["A3"] = "=Bob!D2"
    summary["B3"] = "=Bob!I42"

    alice = workbook.create_sheet("Alice")
    alice["D2"] = "Alice Example"
    alice["E2"] = "alice@example.test"
    alice["F2"] = "090-1234-5678"
    alice["I42"] = 32
    alice["D2"].comment = Comment("Original note about Alice Example", "tester")

    bob = workbook.create_sheet("Bob")
    bob["D2"] = "Bob Example"
    bob["E2"] = "bob@example.test"
    bob["F2"] = "080-0000-1111"
    bob["I42"] = 24

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def make_edge_case_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Name"
    summary["B1"] = "Hours"
    summary["A2"] = "='Carol West'!D2"
    summary["B2"] = "='Carol West'!I42"
    summary["A3"] = "='Carol West Copy'!D2"
    summary["B3"] = "='Carol West Copy'!I42"
    summary["A4"] = "=Delta!D2"
    summary["B4"] = "=Delta!I42"

    carol = workbook.create_sheet("Carol West")
    carol["D2"] = "Carol Example"
    carol["E2"] = "carol@example.test"
    carol["I42"] = 40

    carol_copy = workbook.create_sheet("Carol West Copy")
    carol_copy["D2"] = "Carol Example"
    carol_copy["E2"] = "carol.copy@example.test"
    carol_copy["I42"] = 12
    carol_copy.sheet_state = "hidden"

    delta = workbook.create_sheet("Delta")
    delta["D2"] = "Delta Example"
    delta["E2"] = "delta@example.test"
    delta["I42"] = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


if __name__ == "__main__":
    make_synthetic_workbook(Path(__file__).with_name("synthetic_timecard.xlsx"))
