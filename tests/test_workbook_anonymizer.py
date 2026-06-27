from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from personal_goal_loop.pii_gate import load_denylist, scan_paths
from personal_goal_loop.workbook_anonymizer import (
    anonymize_workbook,
    load_rules,
    sha256_file,
    write_report,
)
from tests.fixtures.make_synthetic_workbook import make_edge_case_workbook, make_synthetic_workbook


ROOT = Path(__file__).resolve().parents[1]


def test_anonymize_workbook_rewrites_names_sheets_formulas_and_metadata(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "synthetic_timecard_anonymized.xlsx"
    make_synthetic_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    rules_path = ROOT / "examples" / "rules.synthetic.json"
    summary = anonymize_workbook(source, output, rules, rules_path=rules_path)

    original = load_workbook(source, data_only=False)
    anonymized = load_workbook(output, data_only=False)

    assert original.sheetnames == ["Summary", "Alice", "Bob"]
    assert anonymized.sheetnames == ["Summary", "P001", "P002"]
    assert anonymized["Summary"]["A2"].value == "=P001!D2"
    assert anonymized["Summary"]["A3"].value == "=P002!D2"
    assert anonymized["P001"]["D2"].value == "P001"
    assert anonymized["P002"]["D2"].value == "P002"
    assert anonymized["P001"]["E2"].value == "[redacted-email]"
    assert anonymized["P001"]["F2"].value == "[redacted-phone]"
    assert anonymized["P001"]["D2"].comment is None
    assert anonymized.properties.creator == "Personal Goal Loop"
    assert summary.renamed_sheets == {"Alice": "P001", "Bob": "P002"}
    assert summary.total_comments_removed == 1
    assert summary.dry_run is False
    assert summary.source_sha256 == sha256_file(source)
    assert summary.output_sha256 == sha256_file(output)
    assert summary.rules_sha256 == sha256_file(rules_path)
    assert summary.policy_version == "workbook-anonymizer/v1"


def test_pii_gate_fails_source_and_passes_anonymized_workbook(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "synthetic_timecard_anonymized.xlsx"
    make_synthetic_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    anonymize_workbook(source, output, rules)

    tokens, patterns = load_denylist(ROOT / "examples" / "denylist.synthetic.json")
    source_findings = list(scan_paths([source], tokens, patterns))
    output_findings = list(scan_paths([output], tokens, patterns))

    assert source_findings
    assert output_findings == []


def test_pii_gate_exclude_patterns_skip_intentional_fixture_files(tmp_path: Path) -> None:
    root = tmp_path / "repo"
    fixture_dir = root / "examples"
    fixture_dir.mkdir(parents=True)
    (fixture_dir / "denylist.synthetic.json").write_text("Alice Example", encoding="utf-8")
    (root / "safe.md").write_text("public template only", encoding="utf-8")

    findings = list(scan_paths([root], ["Alice Example"], []))
    excluded_findings = list(
        scan_paths([root], ["Alice Example"], [], exclude_patterns=["examples/**"])
    )

    assert findings
    assert excluded_findings == []


def test_report_shape_is_json_serializable(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "synthetic_timecard_anonymized.xlsx"
    make_synthetic_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    summary = anonymize_workbook(source, output, rules)

    payload = summary.to_dict()
    assert json.loads(json.dumps(payload))["total_cells_changed"] >= 8
    assert payload["sensitive_fields_redacted"] is True
    assert payload["renamed_sheets"] == {"source_sheet_001": "P001", "source_sheet_002": "P002"}
    assert "Alice" not in json.dumps(payload)

    sensitive_payload = summary.to_dict(include_sensitive=True)
    assert sensitive_payload["sensitive_fields_redacted"] is False
    assert sensitive_payload["renamed_sheets"] == {"Alice": "P001", "Bob": "P002"}


def test_dry_run_writes_no_output_but_returns_receipt(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "dry_run_output.xlsx"
    make_synthetic_workbook(source)

    rules_path = ROOT / "examples" / "rules.synthetic.json"
    rules = load_rules(rules_path)
    summary = anonymize_workbook(source, output, rules, dry_run=True, rules_path=rules_path)

    assert output.exists() is False
    assert summary.dry_run is True
    assert summary.output_sha256 is None
    assert summary.source_sha256 == sha256_file(source)
    assert summary.rules_sha256 == sha256_file(rules_path)
    assert summary.renamed_sheets == {"Alice": "P001", "Bob": "P002"}


def test_edge_case_workbook_handles_quoted_formulas_hidden_sheets_and_duplicate_names(
    tmp_path: Path,
) -> None:
    source = tmp_path / "edge_case_timecard.xlsx"
    output = tmp_path / "edge_case_timecard_anonymized.xlsx"
    make_edge_case_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    summary = anonymize_workbook(source, output, rules)
    anonymized = load_workbook(output, data_only=False)

    assert anonymized.sheetnames == ["Summary", "P001", "P001_2", "P002"]
    assert anonymized["Summary"]["A2"].value == "='P001'!D2"
    assert anonymized["Summary"]["A3"].value == "='P001_2'!D2"
    assert anonymized["Summary"]["A4"].value == "=P002!D2"
    assert anonymized["P001"]["D2"].value == "P001"
    assert anonymized["P001_2"]["D2"].value == "P001"
    assert anonymized["P001_2"].sheet_state == "hidden"
    assert summary.renamed_sheets == {
        "Carol West": "P001",
        "Carol West Copy": "P001_2",
        "Delta": "P002",
    }


def test_golden_summary_subset_is_stable(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "synthetic_timecard_anonymized.xlsx"
    make_synthetic_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    summary = anonymize_workbook(source, output, rules)

    stable_subset = {
        "dry_run": summary.dry_run,
        "policy_version": summary.policy_version,
        "sensitive_fields_redacted": summary.to_dict()["sensitive_fields_redacted"],
        "total_cells_changed": summary.total_cells_changed,
        "total_comments_removed": summary.total_comments_removed,
        "total_hyperlinks_removed": summary.total_hyperlinks_removed,
        "derived_text_replacements_count": summary.derived_text_replacements_count,
        "renamed_sheets_count": summary.to_dict()["renamed_sheets_count"],
        "renamed_sheets": summary.to_dict()["renamed_sheets"],
        "sheets": summary.to_dict()["sheets"],
    }

    expected = json.loads(
        (ROOT / "tests" / "fixtures" / "expected_synthetic_summary.json").read_text(
            encoding="utf-8"
        )
    )
    assert stable_subset == expected


def test_default_report_is_safe_for_pii_gate(tmp_path: Path) -> None:
    source = tmp_path / "synthetic_timecard.xlsx"
    output = tmp_path / "synthetic_timecard_anonymized.xlsx"
    report = tmp_path / "synthetic_timecard_anonymized.anonymize-report.json"
    make_synthetic_workbook(source)

    rules = load_rules(ROOT / "examples" / "rules.synthetic.json")
    summary = anonymize_workbook(source, output, rules)
    write_report(summary, report)

    tokens, patterns = load_denylist(ROOT / "examples" / "denylist.synthetic.json")
    assert list(scan_paths([report], tokens, patterns)) == []
